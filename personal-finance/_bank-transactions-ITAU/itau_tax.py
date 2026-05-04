#!/usr/bin/env python3
"""
itau_tax.py — Brazilian Itaú bank statement → Excel tax report
Usage: python itau_tax.py <input.txt> [output.xlsx] [config.json]

Input format (semicolon-delimited, Itaú export):
  DD/MM/YYYY;DESCRIPTION;AMOUNT
  e.g. 03/01/2025;PIX TRANSF  BANCO I03/01;4459,77

Generates 4-tab Excel:
  1. Resumo Mensal    — categories × months with subtotals
  2. Receitas por Mês — monthly income detail
  3. US Tax Summary   — annual totals for Schedule E / Form 1116
  4. Transações       — raw data with auto-filter

Config file (config.json) controls:
  - Category rules (keyword → category mapping)
  - Category display order
  - US tax notes per category
  - USD/BRL exchange rate for the year
"""

import sys
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── defaults ──────────────────────────────────────────────────────────────────

DEFAULT_CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'config.json')

MONTH_NAMES = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
               7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

# ── colors & styles ───────────────────────────────────────────────────────────

DARK_BLUE  = '1F4E79'
MED_BLUE   = '2E75B6'
BLUE_2F    = '2F5597'
SUBTOT_BLU = 'BDD7EE'
D9E1F2     = 'D9E1F2'
GREEN_375  = '375623'
GREEN_BG   = 'E2EFDA'
YELLOW_BG  = 'FFF2CC'
RED_BG     = 'FCE4D6'
WHITE      = 'FFFFFF'

def _fill(c): return PatternFill('solid', start_color=c)
def _font(bold=False, color='000000', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
_thin = Side(style='thin', color='CCCCCC')
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BRL     = '#,##0.00'
BRL_NEG = '#,##0.00;[RED]-#,##0.00'

def _hdr(cell, bg=DARK_BLUE, fg=WHITE, bold=True, size=10, align='center'):
    cell.fill = _fill(bg); cell.font = _font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = _bdr

def _data(cell, fmt=None, bold=False, bg=None, fg='000000'):
    cell.font = _font(bold=bold, color=fg)
    cell.border = _bdr
    cell.alignment = Alignment(vertical='center')
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = _fill(bg)

def _banner(ws, row, label, ncols, bg=MED_BLUE):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(bg); c.font = _font(bold=True, color=WHITE, size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 16

# ── parse ──────────────────────────────────────────────────────────────────────

def parse_file(path, tax_year):
    records = []
    skipped = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line: continue
            parts = line.split(';')
            if len(parts) != 3:
                skipped.append(line); continue
            date_str, desc, amt_str = parts
            try:
                date = datetime.strptime(date_str, '%d/%m/%Y')
            except ValueError:
                skipped.append(line); continue
            if date.year != tax_year:
                skipped.append(line); continue
            amt = float(amt_str.replace(',', '.'))
            mn = date.month
            records.append({
                'date': date, 'month_num': mn,
                'month': f"{MONTH_NAMES[mn]}/{tax_year}",
                'desc': desc.strip(), 'amount': amt,
                'type': 'Receita' if amt > 0 else 'Despesa'
            })
    if skipped:
        print(f"  ⚠  Skipped {len(skipped)} line(s) (wrong year or bad format)")
    return records

# ── categorize ────────────────────────────────────────────────────────────────

def build_categorizer(rules):
    """Returns a function that maps description → category."""
    compiled = [(cat, [kw.upper() for kw in keywords])
                for cat, keywords in rules.items()]
    def categorize(desc):
        d = desc.upper()
        for cat, keywords in compiled:
            if any(kw in d for kw in keywords):
                return cat
        return 'Outros'
    return categorize

# ── sheet builders ────────────────────────────────────────────────────────────

def build_resumo_mensal(wb, records, cfg):
    ws = wb.active
    ws.title = 'Resumo Mensal'
    NM = 12
    NCOLS = NM + 2  # A + 12 months + total

    tax_year = cfg['tax_year']
    month_labels = [f"{MONTH_NAMES[m]}/{tax_year}" for m in range(1, 13)]
    months_order = list(range(1, 13))

    income_order  = cfg['category_order_income']
    expense_order = cfg['category_order_expense']

    def mval(cat, mn):
        return sum(r['amount'] for r in records if r['category']==cat and r['month_num']==mn)

    # Title
    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    t = ws['A1']
    t.value = f'RESUMO MENSAL — {cfg["account_label"]}  |  {tax_year}'
    t.fill = _fill(DARK_BLUE); t.font = _font(bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Column headers
    _hdr(ws.cell(row=2, column=1, value='Categoria'), align='left')
    ws.column_dimensions['A'].width = 36
    for i, ml in enumerate(month_labels, 2):
        _hdr(ws.cell(row=2, column=i, value=ml))
        ws.column_dimensions[get_column_letter(i)].width = 11
    _hdr(ws.cell(row=2, column=NM+2, value='TOTAL ANUAL'))
    ws.column_dimensions[get_column_letter(NM+2)].width = 14
    ws.row_dimensions[2].height = 20

    r = 3

    def write_cat_row(row, cat, row_bg=None):
        c = ws.cell(row=row, column=1, value=cat)
        c.border = _bdr; c.font = _font()
        c.alignment = Alignment(vertical='center')
        if row_bg: c.fill = _fill(row_bg)
        for mi, mn in enumerate(months_order, 2):
            val = mval(cat, mn)
            cell = ws.cell(row=row, column=mi, value=val if val != 0 else None)
            cell.number_format = BRL_NEG; cell.border = _bdr; cell.font = _font()
            if row_bg: cell.fill = _fill(row_bg)
        tot = ws.cell(row=row, column=NM+2,
                      value=f'=SUM(B{row}:{get_column_letter(NM+1)}{row})')
        tot.number_format = BRL_NEG; tot.border = _bdr; tot.font = _font(bold=True)
        if row_bg: tot.fill = _fill(row_bg)

    def fill_row_bg(row, bg, ncols=NCOLS):
        for col in range(1, ncols+1):
            ws.cell(row=row, column=col).fill = _fill(bg)
            ws.cell(row=row, column=col).border = _bdr

    def subtot_row(row, label, start, end, bg=SUBTOT_BLU, fg='000000'):
        fill_row_bg(row, bg)
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=fg)
        ws.cell(row=row, column=1).fill = _fill(bg)
        for mi in range(2, NM+3):
            cl = get_column_letter(mi)
            f = ws.cell(row=row, column=mi, value=f'=SUM({cl}{start}:{cl}{end})')
            f.number_format = BRL_NEG; f.fill = _fill(bg)
            f.font = _font(bold=True); f.border = _bdr
        return row

    # ---- RECEITAS ----
    _banner(ws, r, '▶  RECEITAS', NCOLS); r += 1
    inc_start = r
    present_income = [c for c in income_order if any(rec['category']==c for rec in records)]
    for cat in present_income:
        write_cat_row(r, cat, row_bg=GREEN_BG); r += 1
    inc_end = r - 1
    total_inc_row = subtot_row(r, 'TOTAL RECEITAS', inc_start, inc_end)
    r += 2

    # ---- DESPESAS ----
    _banner(ws, r, '▶  DESPESAS', NCOLS); r += 1

    # Deductibles
    _banner(ws, r, '   📋  Dedutíveis EUA — Impostos e Condomínio', NCOLS, bg=GREEN_375); r += 1
    ded_start = r
    ded_cats = ['Impostos Governo (IPTU/Taxas)', 'Condomínio']
    for cat in ded_cats:
        if any(rec['category']==cat for rec in records):
            bg = YELLOW_BG if 'Imposto' in cat else GREEN_BG
            write_cat_row(r, cat, row_bg=bg); r += 1
    ded_end = r - 1
    subtot_ded_row = subtot_row(r, '  Subtotal Dedutíveis', ded_start, ded_end, bg=D9E1F2, fg=GREEN_375)
    r += 1

    # Other expenses
    _banner(ws, r, '   📋  Outras Despesas Operacionais', NCOLS, bg=BLUE_2F); r += 1
    other_start = r
    other_cats = [c for c in expense_order if c not in ded_cats]
    for cat in other_cats:
        if any(rec['category']==cat for rec in records):
            write_cat_row(r, cat); r += 1
    other_end = r - 1

    # Total despesas
    fill_row_bg(r, SUBTOT_BLU)
    ws.cell(row=r, column=1, value='TOTAL DESPESAS').font = _font(bold=True)
    ws.cell(row=r, column=1).fill = _fill(SUBTOT_BLU)
    for mi in range(2, NM+3):
        cl = get_column_letter(mi)
        f = ws.cell(row=r, column=mi,
            value=f'=SUM({cl}{ded_start}:{cl}{ded_end})+SUM({cl}{other_start}:{cl}{other_end})')
        f.number_format = BRL_NEG; f.fill = _fill(SUBTOT_BLU)
        f.font = _font(bold=True); f.border = _bdr
    total_exp_row = r; r += 2

    # Net
    fill_row_bg(r, DARK_BLUE)
    ws.cell(row=r, column=1, value='SALDO LÍQUIDO').font = _font(bold=True, color=WHITE, size=11)
    ws.cell(row=r, column=1).fill = _fill(DARK_BLUE)
    for mi in range(2, NM+3):
        cl = get_column_letter(mi)
        f = ws.cell(row=r, column=mi,
            value=f'={cl}{total_inc_row}+{cl}{total_exp_row}')
        f.number_format = BRL_NEG; f.fill = _fill(DARK_BLUE)
        f.font = _font(bold=True, color=WHITE, size=11); f.border = _bdr

    ws.freeze_panes = 'B3'


def build_receitas_mes(wb, records, cfg):
    ws = wb.create_sheet('Receitas por Mês')
    tax_year = cfg['tax_year']

    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value = f'RECEITAS POR MÊS — {tax_year}'
    t.fill = _fill(DARK_BLUE); t.font = _font(bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for i, (h, w) in enumerate(zip(['Mês','Data','Descrição','Categoria','Valor (R$)'],
                                    [10, 14, 48, 28, 16]), 1):
        _hdr(ws.cell(row=2, column=i, value=h))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 18

    income_records = sorted([r for r in records if r['amount'] > 0],
                             key=lambda x: (x['month_num'], x['date']))
    r2 = 3
    for mn in range(1, 13):
        month_recs = [rec for rec in income_records if rec['month_num'] == mn]
        if not month_recs: continue
        start_r = r2
        for rec in month_recs:
            for col in range(1, 6):
                ws.cell(row=r2, column=col).fill = _fill(GREEN_BG)
                ws.cell(row=r2, column=col).border = _bdr
                ws.cell(row=r2, column=col).font = _font()
            ws.cell(row=r2, column=1, value=rec['month'])
            d = ws.cell(row=r2, column=2, value=rec['date'])
            d.number_format = 'DD/MM/YYYY'
            ws.cell(row=r2, column=3, value=rec['desc'])
            ws.cell(row=r2, column=4, value=rec['category'])
            ws.cell(row=r2, column=5, value=rec['amount']).number_format = BRL
            r2 += 1
        end_r = r2 - 1
        for col in range(1, 5):
            ws.cell(row=r2, column=col).fill = _fill(SUBTOT_BLU)
            ws.cell(row=r2, column=col).border = _bdr
        ws.cell(row=r2, column=1, value=f'Subtotal {MONTH_NAMES[mn]}').font = _font(bold=True)
        ws.cell(row=r2, column=1).fill = _fill(SUBTOT_BLU)
        sub = ws.cell(row=r2, column=5, value=f'=SUM(E{start_r}:E{end_r})')
        sub.number_format = BRL; sub.fill = _fill(SUBTOT_BLU)
        sub.font = _font(bold=True); sub.border = _bdr
        r2 += 2

    # Annual total
    for col in range(1, 5):
        ws.cell(row=r2, column=col).fill = _fill(DARK_BLUE)
        ws.cell(row=r2, column=col).border = _bdr
    ws.cell(row=r2, column=1, value=f'TOTAL RECEITAS {tax_year}').font = _font(bold=True, color=WHITE, size=11)
    ws.cell(row=r2, column=1).fill = _fill(DARK_BLUE)
    c = ws.cell(row=r2, column=5, value=f'=SUMIF(E3:E{r2-1},">"&0)')
    c.number_format = BRL; c.fill = _fill(DARK_BLUE)
    c.font = _font(bold=True, color=WHITE, size=11); c.border = _bdr
    ws.freeze_panes = 'A3'


def build_us_tax(wb, records, cfg):
    ws = wb.create_sheet('US Tax Summary')
    tax_year = cfg['tax_year']
    usd_rate = cfg.get('usd_brl_annual_avg')
    notes = cfg.get('us_tax_notes', {})
    income_order  = cfg['category_order_income']
    expense_order = cfg['category_order_expense']

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30

    ws.merge_cells('A1:C1')
    t = ws['A1']
    t.value = f'U.S. TAX SUMMARY — BRAZIL PROPERTY EXPENSES {tax_year}'
    t.fill = _fill(DARK_BLUE); t.font = _font(bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:C2')
    s = ws['A2']
    rate_str = f'  |  USD/BRL rate: {usd_rate}' if usd_rate else '  |  Set USD/BRL rate in config.json'
    s.value = f'{cfg["account_label"]}{rate_str}'
    s.fill = _fill(MED_BLUE); s.font = _font(color=WHITE, size=10)
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    for i, h in enumerate(['Category', 'Annual Total (R$)', 'Notes'], 1):
        _hdr(ws.cell(row=3, column=i, value=h))
    ws.row_dimensions[3].height = 18

    r3 = 4

    def section(row, title, bg=MED_BLUE):
        ws.merge_cells(f'A{row}:C{row}')
        c = ws.cell(row=row, column=1, value=title)
        c.fill = _fill(bg); c.font = _font(bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 16

    def cat_row(row, cat, note='', bg=None):
        total = sum(rec['amount'] for rec in records if rec['category']==cat)
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = _bdr
            ws.cell(row=row, column=col).font = _font()
            if bg: ws.cell(row=row, column=col).fill = _fill(bg)
        ws.cell(row=row, column=1, value=cat)
        c2 = ws.cell(row=row, column=2, value=total)
        c2.number_format = BRL_NEG; c2.font = _font(bold=True)
        if bg: c2.fill = _fill(bg)
        ws.cell(row=row, column=3, value=note).font = _font(color='595959', size=9)
        return row + 1

    def subtot(row, label, start, end, bg=SUBTOT_BLU, fg='000000'):
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = _fill(bg)
            ws.cell(row=row, column=col).border = _bdr
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, color=fg)
        ws.cell(row=row, column=1).fill = _fill(bg)
        f = ws.cell(row=row, column=2, value=f'=SUM(B{start}:B{end})')
        f.number_format = BRL_NEG; f.fill = _fill(bg); f.font = _font(bold=True); f.border = _bdr
        return row

    # Income
    section(r3, '▶  INCOME'); r3 += 1
    inc_s = r3
    present_income = [c for c in income_order if any(rec['category']==c for rec in records)]
    for cat in present_income:
        r3 = cat_row(r3, cat, 'Rental income / transfers', bg=GREEN_BG)
    inc_e = r3 - 1
    ti_row = subtot(r3, 'TOTAL INCOME', inc_s, inc_e); r3 += 2

    # Deductibles
    section(r3, '▶  DEDUCTIBLE EXPENSES  (Schedule E / Form 1116)', bg=GREEN_375); r3 += 1
    ded_s = r3
    r3 = cat_row(r3, 'Impostos Governo (IPTU/Taxas)',
                 notes.get('Impostos Governo (IPTU/Taxas)', 'Real estate taxes'), bg=YELLOW_BG)
    r3 = cat_row(r3, 'Condomínio',
                 notes.get('Condomínio', 'HOA / condo fees'), bg=GREEN_BG)
    ded_e = r3 - 1
    sd_row = subtot(r3, 'Subtotal Deductible', ded_s, ded_e, bg=D9E1F2, fg=GREEN_375); r3 += 2

    # Other expenses
    section(r3, '▶  OTHER OPERATING EXPENSES', bg=BLUE_2F); r3 += 1
    oth_s = r3
    skip = {'Impostos Governo (IPTU/Taxas)', 'Condomínio'}
    for cat in expense_order:
        if cat in skip: continue
        if any(rec['category']==cat for rec in records):
            r3 = cat_row(r3, cat, notes.get(cat, ''))
    oth_e = r3 - 1
    so_row = subtot(r3, 'Subtotal Other Expenses', oth_s, oth_e); r3 += 2

    # Summary
    section(r3, '▶  SUMMARY', bg=DARK_BLUE); r3 += 1
    for label, formula, note in [
        ('TOTAL INCOME',   f'=B{ti_row}', ''),
        ('TOTAL EXPENSES', f'=B{sd_row}+B{so_row}', '(all categories)'),
        ('NET',            f'=B{ti_row}+B{sd_row}+B{so_row}', 'Positive = net income'),
    ]:
        for col in range(1, 4):
            ws.cell(row=r3, column=col).fill = _fill(DARK_BLUE)
            ws.cell(row=r3, column=col).border = _bdr
        ws.cell(row=r3, column=1, value=label).font = _font(bold=True, color=WHITE, size=11)
        ws.cell(row=r3, column=1).fill = _fill(DARK_BLUE)
        gf = ws.cell(row=r3, column=2, value=formula)
        gf.number_format = BRL_NEG; gf.fill = _fill(DARK_BLUE)
        gf.font = _font(bold=True, color=WHITE, size=11); gf.border = _bdr
        ws.cell(row=r3, column=3, value=note).fill = _fill(DARK_BLUE)
        ws.cell(row=r3, column=3).font = _font(color='AAAAAA', size=9); ws.cell(row=r3, column=3).border = _bdr
        r3 += 1

    r3 += 1
    ws.merge_cells(f'A{r3}:C{r3}')
    rate_note = f'USD/BRL annual avg rate: {usd_rate}. ' if usd_rate else 'Set usd_brl_annual_avg in config.json. '
    disc = ws.cell(row=r3, column=1,
        value=f'⚠  {rate_note}All amounts in BRL. Convert to USD for Schedule E filing. Consult a tax advisor.')
    disc.font = Font(name='Arial', italic=True, color='595959', size=9)
    disc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r3].height = 28
    ws.freeze_panes = 'A4'


def build_transacoes(wb, records):
    ws = wb.create_sheet('Transações')
    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value = 'TODAS AS TRANSAÇÕES'
    t.fill = _fill(DARK_BLUE); t.font = _font(bold=True, color=WHITE, size=12)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    for i, (h, w) in enumerate(zip(['Data','Descrição','Categoria','Tipo','Valor (R$)'],
                                    [14, 48, 36, 12, 16]), 1):
        _hdr(ws.cell(row=2, column=i, value=h))
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, rec in enumerate(records, 3):
        bg = GREEN_BG if rec['amount'] > 0 else RED_BG
        for col in range(1, 6):
            ws.cell(row=ri, column=col).fill = _fill(bg)
            ws.cell(row=ri, column=col).font = _font()
            ws.cell(row=ri, column=col).border = _bdr
        ws.cell(row=ri, column=1, value=rec['date']).number_format = 'DD/MM/YYYY'
        ws.cell(row=ri, column=2, value=rec['desc'])
        ws.cell(row=ri, column=3, value=rec['category'])
        ws.cell(row=ri, column=4, value=rec['type'])
        ws.cell(row=ri, column=5, value=rec['amount']).number_format = BRL_NEG

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:E{len(records)+2}'


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path  = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else input_path.replace('.txt', '_tax.xlsx')
    config_path = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_CONFIG_PATH

    print(f"  Input:  {input_path}")
    print(f"  Config: {config_path}")
    print(f"  Output: {output_path}")

    with open(config_path, encoding='utf-8') as f:
        cfg = json.load(f)

    tax_year   = cfg['tax_year']
    categorize = build_categorizer(cfg['category_rules'])

    records = parse_file(input_path, tax_year)
    print(f"  Parsed {len(records)} transactions for {tax_year}")

    for rec in records:
        rec['category'] = categorize(rec['desc'])

    uncategorized = [r for r in records if r['category'] == 'Outros']
    if uncategorized:
        print(f"  ⚠  {len(uncategorized)} transaction(s) landed in 'Outros' — review config.json rules:")
        for r in uncategorized:
            print(f"     {r['date'].strftime('%d/%m/%Y')}  {r['desc'][:50]}  {r['amount']:.2f}")

    wb = openpyxl.Workbook()
    build_resumo_mensal(wb, records, cfg)
    build_receitas_mes(wb, records, cfg)
    build_us_tax(wb, records, cfg)
    build_transacoes(wb, records)

    wb.save(output_path)
    print(f"  ✅  Saved: {output_path}")


if __name__ == '__main__':
    main()
